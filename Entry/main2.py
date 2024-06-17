import datetime
import time

import openpyxl
import pandas as pd
import xlwings as xw
from flask import Flask, render_template, request

defdate = datetime.datetime.today().strftime("%d-%m-%Y")
global file
file = "feescopy1.xlsx"
cols = [1, 2, 3, 7, 13, 11]
out = pd.read_excel(io=file, sheet_name="Student Master", converters={"Cell": str, "Adm": str}, usecols=cols)
app = Flask(__name__)


@app.route("/")
def home():
    return render_template("div.html", defdate=defdate)


@app.route("/", methods=["POST"])
def entry():
    global file
    global A
    wb = openpyxl.load_workbook(file)
    sheet = wb["Daily Fees Entry"]

    for row in sheet.iter_rows(min_col=1, max_col=1):
        if row[0].value is None:
            A = str(row[0].row)
            break
    else:
        A = str(sheet.max_row + 1)

    # B=request.form['Date']
    # C=int(request.form['Receipt'])
    # D=int(request.form['Adm'])
    # E=str(request.form['FeeType'])
    # F=int(request.form['FeeAmt'])

    G = request.values
    response = G.getlist("Receipt")
    # sheet['A'+A]=int(A)+1
    # sheet['B'+A]=B
    # sheet['C'+A]=C
    # sheet['D'+A]=D
    # sheet['E'+A]=E
    # sheet['F'+A]=F

    # temp=out[out['Adm'].to_numpy() == str(D)]
    # temp.reset_index(inplace=True,drop=True)
    # dicdef=pd.DataFrame([{"S.N.":A,"Date":B,"Receipt No.":C,"Fee Type":E,"Fee Amount":F}])
    # result = pd.concat([dicdef, temp], axis=1)
    # outtable=temp.to_html(col_space='50px',classes='data',header='True')
    # wb.save(filename=file)
    # wb.close()
    print(G)
    outtable = None
    return render_template("div.html", defdate=defdate, outtable=outtable)


# @app.route('/show', methods=['GET'])
# def show():
#     start1=time.time()
#     global file
#     global A
#     cols=[1,2,3,7,13,11]

#     with xw.Book("feescopy1.xlsx", mode="r") as book:
#         sheet1 = book.sheets["Daily Fees Entry"]
#         start="A"+A
#         end="F"+A
#         data = sheet1[start+":"+end].options("df", index=False,ndim=2).value
#     showtable=data.to_html(col_space='50px',classes='data',header='True')
#     print(start1-time.time())
#     return render_template("show.html",showtable=showtable)

if __name__ == "__main__":
    app.run()
